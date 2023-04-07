import glob
import os
import smtplib
from email.mime.application import MIMEApplication
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText
from PySide6.QtWidgets import QApplication, QMainWindow, QMessageBox, QFileDialog
from openpyxl import load_workbook
from main_w import Ui_MainWindow

class MainWindow(QMainWindow, Ui_MainWindow):
    def __init__(self, *args, obj=None, **kwargs):
        super(MainWindow, self).__init__(*args, **kwargs)
        self.setupUi(self)
        self.setWindowTitle("거래처 매일 전송")
        self.list_btn.clicked.connect(self.fileopen)
        self.attach_btn.clicked.connect(self.openurl)
        self.send_btn.clicked.connect(self.sendmail)
        self.list_edit.setReadOnly(True)
        self.attach_edit.setReadOnly(True)

    def openurl(self):
        attach_url = QFileDialog.getExistingDirectory(self, 'Select Directory')
        self.attach_edit.setText(attach_url)
        return

    def fileopen(self):
        fname = QFileDialog.getOpenFileName(self, 'Select File', '', 'Excel Files (*.xlsx)')
        self.list_edit.setText(fname[0])
        return

    def sendmail(self):
        mail_id = self.id_edit.text()
        mail_pass = self.pass_edit.text()
        sender = self.sender_edit.text()
        attach_url = self.attach_edit.text()
        list_name = self.list_edit.text()
        subject = self.subject_edit.text()
        body = self.body_edit.toPlainText()

        if mail_id == "":
            QMessageBox.information(self, 'Notice', '이메일을 입력하세요.')

        elif mail_pass == "":
            QMessageBox.information(self, 'Notice', '패스워드를 입력하세요.')

        elif attach_url == "":
            QMessageBox.information(self, 'Notice', '첨부파일 경로를 선택하세요.')

        elif list_name == "":
            QMessageBox.information(self, 'Notice', '거래처 메일 리스트를 선택하세요.')

        else:
            self.send_btn.setText("전송중")
            self.send_btn.setDisabled(True)

            xlsx = load_workbook(list_name, read_only=True)
            sheet = xlsx["거래처명"]

            for row in sheet.iter_rows():
                name = row[0].value
                mail = row[1].value

                if name == None:
                    break

                try:
                    smtp_hwmail = smtplib.SMTP('smtps.hiworks.com', 587)
                    smtp_hwmail.ehlo()
                    # smtp_hwmail.starttls #465
                    smtp_hwmail.login(mail_id, mail_pass)

                except smtplib.SMTPRecipientsRefused:
                    QMessageBox.information(self, 'Notice', '로그인 에러 - 잘 못된 아이디 패스워드')

                except smtplib.SMTPAuthenticationError:
                    QMessageBox.information(self, 'Notice', '로그인 에러')

                except smtplib.SMTPException:
                    QMessageBox.information(self, 'Notice', '에러')

                
                #20230406 : 메일 제목에 업체명 추가
                msg = MIMEMultipart()
                msg['Subject'] = name + " - " + subject
                msg['From'] = sender
                msg['To'] = mail
                msg['Bcc'] = sender

                mail_body = MIMEText(body % name, 'html', 'utf-8')
                msg.attach(mail_body)


                #20230405 : 파일명 앞뒤에 _ 가 들어 가도록 수정
                pdf_list = glob.glob(attach_url + "//*_" + name + "_*.pdf")
                for file_path in pdf_list:
                    with open(file_path, "rb") as f:
                        attach = MIMEApplication(f.read())
                        attach.add_header('Content-Disposition', 'attachment', filename=os.path.split(file_path)[1])
                        print(os.path.split(file_path)[1] + " : 첨부")
                        msg.attach(attach)

                #20230405 : 파일명 앞뒤에 _ 가 들어 가도록 수정
                xl_list = glob.glob(attach_url + "//*_" + name + "_*.xlsx")
                for file_path in xl_list:
                    with open(file_path, "rb") as f:
                        attach = MIMEApplication(f.read())
                        attach.add_header('Content-Disposition', 'attachment', filename=os.path.split(file_path)[1])
                        print(os.path.split(file_path)[1] + " : 첨부")
                        msg.attach(attach)

                smtp_hwmail.send_message(msg)
                print(name + ":전송")
                smtp_hwmail.quit()


            print("완료")
        self.send_btn.setText("완료")
        QMessageBox.information(self, 'Notice', '전송 완료')
        self.send_btn.setDisabled(False)
        self.send_btn.setText("메일 전송")
        self.list_edit.setText("")
        self.attach_edit.setText("")








app = QApplication()
window = MainWindow()
window.show()
app.exec_()